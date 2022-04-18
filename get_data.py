import json
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
import os

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)


# GET DATA
def get_data(url, championship):
    other_teams = []
    home_teams = []
    driver.get(url)

    # Get Home team and Away team
    teams = driver.find_elements(By.CSS_SELECTOR, "tbody tr td a")
    for i, team in enumerate(teams):
        if i % 3 == 0:
            home_teams.append(team.text)
        else:
            other_teams.append(team.text)
    away_teams = [team for i, team in enumerate(other_teams) if i % 2 == 0]

    # Get Average Stats
    home_stats = []
    away_stats = []
    all_tr = driver.find_elements(By.CSS_SELECTOR, "tbody tr")
    for tr in all_tr[2:-1]:
        all_td = tr.find_elements(By.CSS_SELECTOR, "td")
        home_stat = all_td[4].text
        away_stat = all_td[9].text
        home_stats.append(home_stat)
        away_stats.append(away_stat)

    write_data_in_json_file(championship=championship, home_teams=home_teams, home_stats=home_stats,
                            away_teams=away_teams, away_stats=away_stats)


def write_data_in_json_file(championship, home_teams, home_stats, away_teams, away_stats, outfile):
    result = {f"{championship}":
                  {"key": "value"},
              {"ke": "val"},

              }

    if not os.path.exists("./data"):
        os.mkdir("./data")

    if not os.path.exists(f"./data/{championship}"):
        os.mkdir(f"./data/{championship}")

    json_object = json.dumps(data, indent=4)
    with open(f"./data/{championship}/{outfile}", "w") as f:
        f.write(json_object)


def write_in_xlsx_file(championship, home_teams, home_stats, away_teams, away_stats):
    # Check if file exist
    if not os.path.exists("./cards_data.xlsx"):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Inutile"
        workbook.save("./cards_data.xlsx")

    workbook = load_workbook("./cards_data.xlsx")

    # Check if sheet exist
    if championship not in workbook.sheetnames:
        worksheet = workbook.create_sheet(f"{championship}")
        worksheet["A1"] = "Equipes Domicile"
        worksheet["B1"] = "Cartons Domicile"
        worksheet["C1"] = "Equipes Exterieur"
        worksheet["D1"] = "Cartons Exterieur"
        workbook.save("./cards_data.xlsx")

    worksheet = workbook[championship]

    # Write Home Teams
    for i, home_team in enumerate(home_teams):
        worksheet[f"A{i + 2}"] = home_team

    # Write Home Cards
    for i, home_stat in enumerate(home_stats):
        worksheet[f"B{i + 2}"] = home_stat

    # Write Away Teams
    for i, away_team in enumerate(away_teams):
        worksheet[f"C{i + 2}"] = away_team

    # Write Away Cards
    for i, away_stat in enumerate(away_stats):
        worksheet[f"D{i + 2}"] = away_stat

    workbook.save("./cards_data.xlsx")
