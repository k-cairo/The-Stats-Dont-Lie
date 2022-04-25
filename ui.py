import os
from datetime import datetime, timedelta, date
from functools import partial
from tkinter import Tk, Button, Label, Text, END
import get_data
import get_iframes
import get_matchs
import show_infos
import json
from constant import CONVERSION_LIST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from subprocess import Popen

today = datetime.today().strftime("%d-%m-%Y")
tomorrow = (date.today() + timedelta(days=1)).strftime("%d-%m-%Y")
j2 = (date.today() + timedelta(days=2)).strftime("%d-%m-%Y")
j3 = (date.today() + timedelta(days=3)).strftime("%d-%m-%Y")
j4 = (date.today() + timedelta(days=4)).strftime("%d-%m-%Y")
j5 = (date.today() + timedelta(days=5)).strftime("%d-%m-%Y")


def define_bet(total_stats):
    if total_stats >= 13.5:
        return "+ 12.5"
    elif total_stats >= 12.5:
        return "+ 11.5"
    elif total_stats >= 11.5:
        return "+ 10.5"
    elif total_stats >= 10.5:
        return "+ 9.5"
    elif total_stats >= 9.5:
        return "+ 8.5"
    elif total_stats >= 8.5:
        return "+ 7.5"
    elif total_stats >= 7.5:
        return "+ 6.5"
    elif total_stats >= 6.5:
        return "+ 5.5"
    elif total_stats >= 5.5:
        return "+ 4.5"
    elif total_stats >= 4.5:
        return "+ 3.5"
    elif total_stats >= 3.5:
        return "+ 2.5"
    elif total_stats >= 2.5:
        return "+ 1.5"
    elif total_stats >= 1.5:
        return "+ 0.5"
    else:
        return "+0"


def championship_convert_name(old_championship, dictionary):
    for key, value in dictionary.items():
        if old_championship == value:
            return key.replace(":", "-")


def write_in_excel_file(date, championship, home_team, away_team, bet_cards, bet_corners, row):
    if not os.path.exists("./Historique"):
        os.mkdir("Historique")

    if not os.path.exists("./Historique/Historique.xlsx"):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Suivi Bets"
        workbook.save("./Historique/Historique.xlsx")

    workbook = load_workbook("./Historique/Historique.xlsx")

    if date not in workbook.sheetnames:
        worksheet = workbook.create_sheet(f"{date}")
        worksheet["A1"] = "Championnats"
        worksheet["B1"] = "Equipes Domicile"
        worksheet["C1"] = "Equipes Exterieur"
        worksheet["D1"] = "Bet Cartons"
        worksheet["E1"] = "Bet Corners"
        worksheet["F1"] = "Cartons réels"
        worksheet["G1"] = "Corners Réels"
        worksheet["H1"] = "Résultats Bet Cartons"
        worksheet["I1"] = "Résultats Bet Corners"
        worksheet["J1"] = "Résultats Bet Global"
        worksheet["K1"] = "Date"

        letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
        for letter in letters:
            worksheet[f"{letter}1"].font = Font(bold=True)

        workbook.save("./Historique/Historique.xlsx")

    worksheet = workbook[date]

    worksheet[f"A{row}"] = championship
    worksheet[f"B{row}"] = home_team
    worksheet[f"C{row}"] = away_team
    worksheet[f"D{row}"] = float(bet_cards.strip("+"))
    worksheet[f"E{row}"] = float(bet_corners.strip("+"))
    worksheet[f"K{row}"] = date

    # Width Cells Auto-adjust
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value

    workbook.save("./Historique/Historique.xlsx")


class UserInterface:
    BUTTONS_WIDTH = 29
    LEFT_RIGHT_BUTTONS_HEIGHT = 8

    def __init__(self):
        self.window = Tk()
        self.window.title("Statistiques Corners & Cartons")
        self.window.geometry("1070x465")

        # Label
        self.main_label = Label(text="Statistiques Corners & Cartons", )
        self.main_label.grid(row=0, column=0, columnspan=5)

        # Left Buttons GET
        self.get_cards_iframes_button = Button(
            text="Get Cards Iframes",
            command=get_iframes.get_all_cards_iframes,
            width=UserInterface.BUTTONS_WIDTH,
            height=UserInterface.LEFT_RIGHT_BUTTONS_HEIGHT)
        self.get_cards_iframes_button.grid(row=1, column=0)

        self.get_corners_iframes_button = Button(
            text="Get Corners Iframes",
            command=get_iframes.get_all_corners_iframes,
            width=UserInterface.BUTTONS_WIDTH,
            height=UserInterface.LEFT_RIGHT_BUTTONS_HEIGHT)
        self.get_corners_iframes_button.grid(row=2, column=0)

        self.get_datas_button = Button(
            text="Get Datas",
            command=self.check_before_get_datas,
            width=UserInterface.BUTTONS_WIDTH,
            height=UserInterface.LEFT_RIGHT_BUTTONS_HEIGHT)
        self.get_datas_button.grid(row=3, column=0)

        # Right Buttons VIEW
        self.view_cards_iframes_button = Button(
            text="Launch Game Analysis",
            command=self.analysis_excel_data,
            width=UserInterface.BUTTONS_WIDTH,
            height=UserInterface.LEFT_RIGHT_BUTTONS_HEIGHT)
        self.view_cards_iframes_button.grid(row=1, column=4)

        self.view_corners_iframes_button = Button(
            text="Launch Excel File",
            command=self.launch_excel_file,
            width=UserInterface.BUTTONS_WIDTH,
            height=UserInterface.LEFT_RIGHT_BUTTONS_HEIGHT)
        self.view_corners_iframes_button.grid(row=2, column=4)

        self.view_datas_button = Button(
            text="View Datas\nComing Next",
            state="disabled",
            width=UserInterface.BUTTONS_WIDTH,
            height=UserInterface.LEFT_RIGHT_BUTTONS_HEIGHT)
        self.view_datas_button.grid(row=3, column=4)

        # Bottom Buttons GET
        self.get_todays_matchs_button = Button(
            text=f"Get Matchs : {datetime.today().strftime('%A %d %B')}",
            command=partial(get_matchs.get_all_matchs, date=today),
            width=UserInterface.BUTTONS_WIDTH)
        self.get_todays_matchs_button.grid(row=4, column=0)

        self.get_tomorrows_matchs_button = Button(
            text=f"Get Matchs : {(datetime.today() + timedelta(days=1)).strftime('%A %d %B')}",
            command=partial(get_matchs.get_all_matchs, date=tomorrow), width=UserInterface.BUTTONS_WIDTH)
        self.get_tomorrows_matchs_button.grid(row=4, column=1)

        self.get_J2_matchs_button = Button(
            text=f"Get Matchs : {(datetime.today() + timedelta(days=2)).strftime('%A %d %B')}",
            command=partial(get_matchs.get_all_matchs, date=j2), width=UserInterface.BUTTONS_WIDTH)
        self.get_J2_matchs_button.grid(row=4, column=2)

        self.get_J3_matchs_button = Button(
            text=f"Get Matchs : {(datetime.today() + timedelta(days=3)).strftime('%A %d %B')}",
            command=partial(get_matchs.get_all_matchs, date=j3), width=UserInterface.BUTTONS_WIDTH)
        self.get_J3_matchs_button.grid(row=4, column=3)

        self.get_J4_matchs_button = Button(
            text=f"Get Matchs : {(datetime.today() + timedelta(days=4)).strftime('%A %d %B')}",
            command=partial(get_matchs.get_all_matchs, date=j4), width=UserInterface.BUTTONS_WIDTH)
        self.get_J4_matchs_button.grid(row=4, column=4)

        # Bottom Buttons VIEW
        self.view_todays_matchs_button = Button(
            text=f"View Matchs : {datetime.today().strftime('%A %d %B')}",
            command=partial(self.read_match_list, date=today),
            width=UserInterface.BUTTONS_WIDTH)
        self.view_todays_matchs_button.grid(row=5, column=0)

        self.view_tomorrows_matchs_button = Button(
            text=f"View Matchs : {(datetime.today() + timedelta(days=1)).strftime('%A %d %B')}",
            command=partial(self.read_match_list, date=tomorrow),
            width=UserInterface.BUTTONS_WIDTH)
        self.view_tomorrows_matchs_button.grid(row=5, column=1)

        self.view_tomorrows_matchs_button = Button(
            text=f"View Matchs : {(datetime.today() + timedelta(days=2)).strftime('%A %d %B')}",
            command=partial(self.read_match_list, date=j2),
            width=UserInterface.BUTTONS_WIDTH)
        self.view_tomorrows_matchs_button.grid(row=5, column=2)

        self.view_tomorrows_matchs_button = Button(
            text=f"View Matchs : {(datetime.today() + timedelta(days=3)).strftime('%A %d %B')}",
            command=partial(self.read_match_list, date=j3),
            width=UserInterface.BUTTONS_WIDTH)
        self.view_tomorrows_matchs_button.grid(row=5, column=3)

        self.view_tomorrows_matchs_button = Button(
            text=f"View Matchs : {(datetime.today() + timedelta(days=4)).strftime('%A %d %B')}",
            command=partial(self.read_match_list, date=j4),
            width=UserInterface.BUTTONS_WIDTH)
        self.view_tomorrows_matchs_button.grid(row=5, column=4)

        # Text
        self.textbox = Text(master=self.window)
        self.textbox.grid(row=1, rowspan=3, column=1, columnspan=3)
        self.window.mainloop()

    def launch_excel_file(self):
        if not os.path.exists("./Historique/Historique.xlsx"):
            self.textbox.insert(index=1.0, chars="File not found\nLaunch 'View Match ...'")
        else:
            Popen(r'C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE ./Historique/Historique.xlsx')

    def read_match_list(self, date):
        excel_row = 1
        self.textbox.delete('1.0', END)
        try:
            with open(f"./Liste de Matchs/{date}.json", "r") as f:
                matchs_list = json.load(f)
        except FileNotFoundError:
            self.textbox.insert(index=1.0, chars="Please get matchs list first")
        else:
            for championship in matchs_list:
                matchs = matchs_list[championship]
                for match in matchs:
                    excel_row += 1
                    match_split = match.split("|")
                    home_team = match_split[0]
                    away_team = match_split[1]
                    # show_infos.debug_format_teams(champ=championship, ht=home_team, at=away_team) # Only for Debugging
                    cards_for_stats = show_infos.get_card_for(champ=championship, ht=home_team, at=away_team)
                    cards_against_stats = show_infos.get_card_against(champ=championship, ht=home_team, at=away_team)
                    corners_for_stats = show_infos.get_corner_for(champ=championship, ht=home_team, at=away_team)
                    corners_against_stats = show_infos.get_corner_against(champ=championship, ht=home_team,
                                                                          at=away_team)
                    total_cards = float(cards_against_stats) + float(cards_for_stats)
                    define_bet(total_stats=total_cards)
                    total_corners = float(corners_against_stats) + float(corners_for_stats)
                    content = f"Championnat : {championship_convert_name(old_championship=championship, dictionary=CONVERSION_LIST)}" \
                              f"\nMatch : {match.replace('|', ' - ')}" \
                              f"\n{cards_for_stats} Cards For" \
                              f"\n{cards_against_stats} Cards Against" \
                              f"\n{corners_for_stats} Corners For" \
                              f"\n{corners_against_stats} Corners Against" \
                              f"\nTotal Cards: {total_cards}" \
                              f"\nTotal Corners: {total_corners}" \
                              f"\nBet à tenter : {define_bet(total_stats=total_cards)} Cards" \
                              f"\nBet à tenter : {define_bet(total_stats=total_corners)} Corners\n\n\n"
                    self.textbox.insert(index=1.0, chars=content)
                    write_in_excel_file(date=date, championship=championship_convert_name(old_championship=championship, dictionary=CONVERSION_LIST), home_team=home_team, away_team=away_team,
                                        bet_cards=define_bet(total_stats=total_cards),
                                        bet_corners=define_bet(total_stats=total_corners), row=excel_row)
            excel_row = 2

    def check_before_get_datas(self):
        for path in get_data.ALL_PATHS:
            if not os.path.exists(path):
                return self.textbox.insert(index=1.0, chars="Please Get Cards Iframes & Corners Iframes First !!!")
            else:
                get_data.get_all_datas()

    def analysis_excel_data(self):
        try:
            workbook = load_workbook("./Historique/Historique.xlsx")
        except FileNotFoundError:
            self.textbox.delete('1.0', END)
            self.textbox.insert(index=1.0, chars="File not found")
        else:
            for sheet in workbook.sheetnames[1:]:
                ws = workbook[sheet]

                row = 2
                while ws[f"A{row}"].value != None:
                    # Cards Logiq
                    if not ws[f"F{row}"].value == None:
                        if ws[f"D{row}"].value < ws[f"F{row}"].value:
                            ws[f"H{row}"] = "Passed"
                        else:
                            ws[f"H{row}"] = "Failed"

                    # Corners Logiq
                    if not ws[f"G{row}"].value == None:
                        if ws[f"E{row}"].value < ws[f"G{row}"].value:
                            ws[f"I{row}"] = "Passed"
                        else:
                            ws[f"I{row}"] = "Failed"

                    # Bet Global Logiq
                    if ws[f"I{row}"].value != None and ws[f"H{row}"].value != None:
                        if ws[f"I{row}"].value == "Passed" and ws[f"H{row}"].value == "Passed":
                            ws[f"J{row}"] = "Passed"
                        else:
                            ws[f"J{row}"] = "Failed"
                    row += 1
                row = 2

            workbook.save("./Historique/Historique.xlsx")

