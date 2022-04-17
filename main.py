import json
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
import os
from constant import CARDS, CORNERS, LEAGUES_URLS, TEAMS_IN_CHAMPIONSHIP

iframes_yc_for = {}
iframes_yc_against = {}

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)


####################################### GET IFRAMES ###################################################################
def get_iframe(link, championship):
    driver.get(link)
    nb_team = TEAMS_IN_CHAMPIONSHIP[championship]
    try:
        iframes = driver.find_elements(By.CSS_SELECTOR, f"div.embed-container{nb_team} iframe")
        iframe_yc_for = iframes[0].get_attribute('src')
        iframe_yc_against = iframes[1].get_attribute('src')
    except IndexError:
        print(f"Erreur avec le championnat : {championship}")
    else:
        iframes_yc_for[championship] = iframe_yc_for
        iframes_yc_against[championship] = iframe_yc_against


def write_in_json_file():
    json_object_1 = json.dumps(iframes_yc_for, indent=4)
    with open("./iframes_YC_for.json", "w") as f:
        f.write(json_object_1)

    json_object_2 = json.dumps(iframes_yc_against, indent=4)
    with open("./iframes_YC_against.json", "w") as f:
        f.write(json_object_2)


def get_all_cards_iframes():
    for key, value in LEAGUES_URLS.items():
        card_link = value + CARDS
        get_iframe(link=card_link, championship=key)
    write_in_json_file()








# GET DATA
def get_data(url, championship):
    other_teams = []
    home_teams = []
    driver.get(url)

    # Get Home team and Away team
    teams = driver.find_elements(By.CSS_SELECTOR, "tbody tr td a")
    for i, team in enumerate(teams):
        if i % 3 == 0:
            home_teams.append(team.text)
        else:
            other_teams.append(team.text)
    away_teams = [team for i, team in enumerate(other_teams) if i % 2 == 0]

    # Get Average Stats
    home_stats = []
    away_stats = []
    all_tr = driver.find_elements(By.CSS_SELECTOR, "tbody tr")
    for tr in all_tr[2:-1]:
        all_td = tr.find_elements(By.CSS_SELECTOR, "td")
        home_stat = all_td[4].text
        away_stat = all_td[9].text
        home_stats.append(home_stat)
        away_stats.append(away_stat)

    write_in_xlsx_file(championship=championship, home_teams=home_teams, home_stats=home_stats, away_teams=away_teams,
                       away_stats=away_stats)


def write_in_xlsx_file(championship, home_teams, home_stats, away_teams, away_stats):
    # Check if file exist
    if not os.path.exists("./cards_data.xlsx"):
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Inutile"
        workbook.save("./cards_data.xlsx")

    workbook = load_workbook("./cards_data.xlsx")

    # Check if sheet exist
    if championship not in workbook.sheetnames:
        worksheet = workbook.create_sheet(f"{championship}")
        worksheet["A1"] = "Equipes Domicile"
        worksheet["B1"] = "Cartons Domicile"
        worksheet["C1"] = "Equipes Exterieur"
        worksheet["D1"] = "Cartons Exterieur"
        workbook.save("./cards_data.xlsx")

    worksheet = workbook[championship]

    # Write Home Teams
    for i, home_team in enumerate(home_teams):
        worksheet[f"A{i + 2}"] = home_team

    # Write Home Cards
    for i, home_stat in enumerate(home_stats):
        worksheet[f"B{i + 2}"] = home_stat

    # Write Away Teams
    for i, away_team in enumerate(away_teams):
        worksheet[f"C{i + 2}"] = away_team

    # Write Away Cards
    for i, away_stat in enumerate(away_stats):
        worksheet[f"D{i + 2}"] = away_stat

    workbook.save("./cards_data.xlsx")


if __name__ == "__main__":
    get_all_cards_iframes()

    # with open("./iframes.json", "r") as f:
    #     content = json.load(f)
    #
    #     for championship, iframe in content.items():
    #         get_data(url=iframe, championship=championship)